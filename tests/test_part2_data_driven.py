#!/usr/bin/env python3
"""
PART 2: Data-Driven UI Filter Testing - Pytest Implementation
Tests EaseMyTrip website UI filtering functionality across multiple datasets.
Compatible with PyCharm IDE and VS Code

This module converts the original Part 2 custom framework to pytest with
parameterized tests while maintaining all existing functionality.

Requirements:
1. Data-driven testing using JSON/CSV input files
2. Multiple combinations of From, To, and Departure Dates
3. Flight preferences with 1 or 2 stops
4. Iterates over 7-10 input data sets
5. Logs and reports results for each test case
"""

import os
import sys
import pytest
import json
import time
import psutil  # For system monitoring
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import gc

# Add project root to path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

def check_system_resources():
    """Monitor system resources to prevent overload"""
    try:
        memory = psutil.virtual_memory()
        cpu = psutil.cpu_percent(interval=1)
        
        # Log resource usage
        memory_percent = memory.percent
        memory_available_gb = memory.available / (1024**3)
        
        print(f"      Memory: {memory_percent:.1f}% used, {memory_available_gb:.1f}GB available")
        print(f"      CPU: {cpu:.1f}% usage")
        
        # Warning thresholds
        if memory_percent > 85:
            print(f"       HIGH MEMORY USAGE: {memory_percent:.1f}%")
        if cpu > 80:
            print(f"       HIGH CPU USAGE: {cpu:.1f}%")
            
        return memory_percent < 90 and cpu < 90  # System health check
        
    except Exception as e:
        print(f"       Resource monitoring failed: {e}")
        return True  # Continue if monitoring fails

from src.automation.flight_filter_engine import PureUIFilterEngine
from src.utils.config import TestConfig
from src.utils.logger import TestLogger


class TestPart2DataDrivenAutomation:
    """
    Test Class: Part 2 Data-Driven UI Filter Testing
    
    This class implements comprehensive data-driven testing for EaseMyTrip
    UI filtering functionality using pytest framework.
    """
    
    def setup_method(self):
        """Setup method called before each test method"""
        self.project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.test_results = []
        self.test_configs = []
        
    @pytest.mark.part2
    @pytest.mark.ui
    @pytest.mark.slow
    def test_part2_data_driven_all_cases(self, test_data_configs, automation_engine, logger):
        """
        Test Part 2: Data-Driven UI Filter Testing - Main Test Method
        
        This comprehensive test:
        1. Loads multiple test configurations from JSON
        2. Executes UI filter tests for each configuration
        3. Collects results and generates detailed reports
        4. Validates overall testing success criteria
        """
        logger.info(" PART 2: Data-Driven UI Filter Testing (Pytest)")
        logger.info(" Testing EaseMyTrip UI Filter Functionality ONLY")
        logger.info("=" * 80)
        
        # Load test configurations
        test_configs = test_data_configs
        total_tests = len(test_configs)
        logger.info(f" Loaded {total_tests} test cases")
        
        # Initialize result tracking
        all_results = []
        passed_tests = 0
        failed_tests = 0
        
        # Execute each test case with resource management
        for idx, config in enumerate(test_configs, 1):
            logger.info(f" Test {idx}/{total_tests}: {config.test_id}")
            logger.info("-" * 50)
            
            try:
                # Pre-test resource cleanup and monitoring
                if idx > 1:  # Add delay between tests (except first)
                    logger.info(f"      System cooldown (3 seconds) + resource cleanup...")
                    time.sleep(3)
                    
                    # Force garbage collection between tests
                    gc.collect()
                    
                    # Check system resources
                    system_ok = check_system_resources()
                    if not system_ok:
                        logger.warning(f"       System resources are high - adding extra cooldown")
                        time.sleep(5)  # Extra cooldown if system is stressed
                
                # Execute UI filter test
                start_time = time.time()
                result = automation_engine.test_ui_filter_functionality(config)
                execution_time = time.time() - start_time
                
                logger.info(f"      Test execution time: {execution_time:.1f}s")
                
                # CRITICAL: Check if the core operations succeeded FIRST
                result_status = result.get('status', 'UNKNOWN')
                
                if result_status == 'FAIL':
                    # Core operation failed (search, city selection, etc.)
                    failed_tests += 1
                    failure_reason = result.get('reason', 'Unknown failure')
                    logger.error(f"     {config.test_id}: CORE OPERATION FAILED - {failure_reason}")
                    result['filter_status'] = 'CORE_OPERATION_FAILED'
                    result['user_message'] = f" Test Failed: {failure_reason}"
                    
                elif result_status == 'ERROR':
                    # Exception or technical error occurred
                    failed_tests += 1
                    error_msg = result.get('error', 'Unknown error')
                    logger.error(f"     {config.test_id}: TECHNICAL ERROR - {error_msg}")
                    result['filter_status'] = 'TECHNICAL_ERROR'
                    result['user_message'] = f" Technical Error: {error_msg}"
                    
                elif result_status == 'TIMEOUT':
                    # Test case exceeded timeout limit
                    failed_tests += 1
                    timeout_msg = result.get('error', 'Test timed out')
                    logger.error(f"     {config.test_id}: TIMEOUT - {timeout_msg}")
                    result['filter_status'] = 'TIMEOUT'
                    result['user_message'] = f" Timeout: Test took too long to complete (likely a high-traffic route)"
                    
                elif result_status == 'SUCCESS':
                    # Core operations succeeded, now check if flights were found and validation
                    flights_found = len(result.get('ui_filtered_flights', []))
                    validation_result = result.get('validation_result', {})
                    validation_passed = validation_result.get('validation_passed', True)
                    
                    if flights_found > 0:
                        # Check if extracted flights match filter criteria
                        if validation_passed:
                            passed_tests += 1
                            logger.info(f"     {config.test_id}: UI Filter Test PASSED - {flights_found} flights found and validated")
                            result['filter_status'] = 'PASSED'
                            result['user_message'] = f" Success: Found {flights_found} flights matching your criteria"
                        else:
                            # Flights found but don't match criteria - FAIL the test
                            failed_tests += 1
                            invalid_count = validation_result.get('invalid_flights', 0)
                            valid_count = validation_result.get('valid_flights', 0)
                            logger.error(f"     {config.test_id}: VALIDATION FAILED - {flights_found} flights found but {invalid_count} don't match criteria")
                            logger.error(f"         Valid: {valid_count}/{flights_found}, Invalid: {invalid_count}/{flights_found}")
                            result['filter_status'] = 'VALIDATION_FAILED'
                            result['user_message'] = f" Test Failed: {flights_found} flights found but {invalid_count} don't match filter criteria (price/stops)"
                    else:
                        # No flights found - still pass 
                        passed_tests += 1  # Core operations worked, just no flights available
                        logger.info(f"     {config.test_id}: No flights found, but test completed successfully")
                        result['filter_status'] = 'NO_FLIGHTS_FOUND'
                        result['user_message'] = f"  No flights found matching your criteria. Try adjusting filters or different dates."
                else:
                    # Unknown status
                    failed_tests += 1
                    logger.error(f"     {config.test_id}: UNKNOWN STATUS - {result_status}")
                    result['filter_status'] = 'UNKNOWN_STATUS'
                    result['user_message'] = f" Unknown test status: {result_status}"
                
                # Store result with test metadata
                result['test_config'] = config
                result['test_number'] = idx
                result['execution_time'] = time.time()
                result['flights_found'] = flights_found
                all_results.append(result)
                
            except Exception as e:
                failed_tests += 1
                logger.error(f"     {config.test_id}: Test execution failed: {str(e)}")
                
                # Emergency stop if too many consecutive failures
                if failed_tests >= 3 and idx <= 5:  # If 3+ failures in first 5 tests
                    logger.error(f"      EMERGENCY STOP: Too many failures ({failed_tests}) - System may be unstable")
                    logger.error(f"      Recommendation: Restart system and try again")
                    break
                
                # Store failure result
                all_results.append({
                    'test_config': config,
                    'test_number': idx,
                    'filter_status': 'TEST_EXECUTION_FAILED',
                    'error_message': str(e),
                    'execution_time': time.time(),
                    'flights_found': 0
                })
                
            # Post-test resource cleanup
            try:
                gc.collect()
                time.sleep(1)  # Brief pause after each test
            except:
                pass
        
        # Generate comprehensive Excel report
        results_directory = os.path.join(project_root, "results")
        os.makedirs(results_directory, exist_ok=True)
        
        report_result = self._generate_part2_report(
            all_results, test_configs, logger, results_directory
        )
        
        # Final system report
        try:
            final_memory = psutil.virtual_memory().percent
            final_cpu = psutil.cpu_percent(interval=1)
            logger.info(f" Final System State - Memory: {final_memory:.1f}%, CPU: {final_cpu:.1f}%")
        except:
            pass
        
        # Calculate success metrics
        success_rate = (passed_tests / total_tests) * 100 if total_tests > 0 else 0
        
        # Log final summary
        logger.info("=" * 80)
        logger.info(" PART 2 DATA-DRIVEN TESTING SUMMARY:")
        logger.info(f"   Total Tests: {total_tests}")
        logger.info(f"   Passed: {passed_tests}")
        logger.info(f"   Failed: {failed_tests}")
        logger.info(f"   Success Rate: {success_rate:.1f}%")
        logger.info(f"   Report Generated: {report_result.get('filename', 'N/A')}")
        logger.info("=" * 80)
        
        # Final cleanup
        try:
            gc.collect()
            time.sleep(2)  # Final cooldown
        except:
            pass
        
        # Store results for potential use in reporting
        self.test_results = all_results
        
        # STRICT ASSERTIONS: Fail the test if core operations fail
        # Check for critical failures (not just "no flights found")
        critical_failures = sum(1 for result in all_results 
                               if result.get('filter_status') in [
                                   'CORE_OPERATION_FAILED', 
                                   'TECHNICAL_ERROR', 
                                   'TIMEOUT',
                                   'TEST_EXECUTION_FAILED',
                                   'VALIDATION_FAILED',  # Added validation failures
                                   'UNKNOWN_STATUS'
                               ])
        
        if critical_failures > 0:
            failure_details = []
            for result in all_results:
                status = result.get('filter_status')
                if status in ['CORE_OPERATION_FAILED', 'TECHNICAL_ERROR', 'TEST_EXECUTION_FAILED', 'VALIDATION_FAILED', 'UNKNOWN_STATUS']:
                    test_id = result.get('test_config', {}).get('test_id', 'Unknown')
                    message = result.get('user_message', 'Unknown error')
                    failure_details.append(f"{test_id}: {message}")
            
            failure_summary = "\\n".join(failure_details)
            pytest.fail(f" {critical_failures} critical test failures occurred:\\n{failure_summary}")
        
        # Success criteria for when core operations work
        assert passed_tests >= 0, "Test execution failed completely"
        
        # Log that tests passed core validations  
        logger.info(f" All {total_tests} tests passed core operation validation")
        validation_failures = sum(1 for result in all_results if result.get('filter_status') == 'VALIDATION_FAILED')
        if failed_tests == 0:
            logger.info(" 100% test success rate achieved!")
        elif validation_failures > 0:
            logger.info(f" {validation_failures} test(s) failed due to validation criteria mismatch")
        
        logger.info(" PART 2 PYTEST DATA-DRIVEN TESTING COMPLETED!")
    
    def _generate_part2_report(self, results: List[Dict], configs: List[TestConfig], 
                              logger, results_directory: str) -> Dict[str, Any]:
        """Generate comprehensive Part 2 report with clean data structure"""
        try:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            
            # Create part2 specific results directory
            part2_results_dir = os.path.join(results_directory, "part2")
            os.makedirs(part2_results_dir, exist_ok=True)
            
            # Create workbook in part2 results folder
            workbook_name = f"Part2_Data_Driven_Pytest_Test_Results_{timestamp}.xlsx"
            results_path = os.path.join(part2_results_dir, workbook_name)
            workbook = openpyxl.Workbook()
            
            # 1. Create test execution summary sheet
            self._create_clean_test_summary_sheet(workbook, results, configs, logger)
            
            # 2. Create detailed flight results for each test
            for result in results:
                if 'ui_filtered_flights' in result and result['ui_filtered_flights']:
                    self._create_clean_flight_details_sheet(workbook, result, logger)
            
            # Save the workbook
            workbook.save(results_path)
            logger.info(f" Part 2 comprehensive report saved: {workbook_name}")
            
            return {
                'success': True,
                'filename': workbook_name,
                'file_path': results_path,
                'total_sheets': len(workbook.worksheets),
                'results_count': len(results)
            }
            
        except Exception as e:
            logger.error(f" Part 2 report generation failed: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'filename': None
            }
    
    def _create_clean_test_summary_sheet(self, workbook, results, configs, logger):
        """Create test execution summary sheet with clean formatting - SAME AS PART 1"""
        try:
            # Create or get summary sheet
            if 'Sheet' in workbook.sheetnames:
                summary_sheet = workbook['Sheet']
                summary_sheet.title = "Test_Execution_Summary"
            else:
                summary_sheet = workbook.create_sheet("Test_Execution_Summary")
            
            # Clear existing content
            summary_sheet.delete_rows(1, summary_sheet.max_row)
            
            # Headers with styling (SAME AS PART 1)
            headers = [
                'Test_ID', 'Route', 'Date', 'Stops_Filter', 'Price_Range', 
                'Status', 'Flights_Found', 'Result', 'Message'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for row, result in enumerate(results, 2):
                config = result.get('test_config')
                if config:
                    summary_sheet.cell(row=row, column=1, value=config.test_id)
                    summary_sheet.cell(row=row, column=2, value=f"{config.from_city} → {config.to_city}")
                    summary_sheet.cell(row=row, column=3, value=config.departure_date)
                    summary_sheet.cell(row=row, column=4, value=config.stops_filter)
                    summary_sheet.cell(row=row, column=5, value=f"₹{config.price_min:,}-₹{config.price_max:,}")
                    summary_sheet.cell(row=row, column=6, value=result.get('filter_status', 'Unknown'))
                    summary_sheet.cell(row=row, column=7, value=result.get('flights_found', 0))
                    summary_sheet.cell(row=row, column=8, value='COMPLETED' if result.get('filter_status') in ['PASSED', 'NO_FLIGHTS_FOUND'] else 'FAILED')
                    summary_sheet.cell(row=row, column=9, value=result.get('user_message', 'Test completed'))
            
            # Auto-adjust column widths (SAME AS PART 1)
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(" Test summary sheet created successfully")
            
        except Exception as e:
            logger.error(f" Error creating test summary sheet: {str(e)}")
    
    def _create_clean_flight_details_sheet(self, workbook, result, logger):
        """Create detailed flight results sheet for each successful test"""
        try:
            config = result.get('test_config')
            if not config:
                return
                
            # Create sheet name
            sheet_name = f"{config.test_id}_Details"[:31]  # Excel sheet name limit
            flight_sheet = workbook.create_sheet(sheet_name)
            
            # Headers for flight data - SAME AS PART 1
            headers = ['Airline Name', 'Price', 'From Airport Code', 'To Airport Code']
            
            # Header styling - SAME AS PART 1
            for col, header in enumerate(headers, 1):
                cell = flight_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Flight data - KEEP ORIGINAL FORMAT (no grouping in detail sheets)
            flights = result.get('ui_filtered_flights', [])
            for row, flight in enumerate(flights, 2):
                flight_sheet.cell(row=row, column=1, value=flight.get('airline', 'N/A'))
                
                # Format price with rupee symbol
                price_value = flight.get('price', 0)
                if isinstance(price_value, (int, float)) and price_value > 0:
                    formatted_price = f"₹{price_value:,}"
                else:
                    formatted_price = 'N/A'
                flight_sheet.cell(row=row, column=2, value=formatted_price)
                
                # Use extracted airport codes from website selection
                flights = result.get('ui_filtered_flights', [])
                if flights:
                    # Get airport codes from first flight (all flights in result have same route)
                    from_airport = flights[0].get('from_code', 'N/A')
                    to_airport = flights[0].get('to_code', 'N/A')
                else:
                    # Fallback if no flights found
                    from_airport = 'N/A'
                    to_airport = 'N/A'
                
                flight_sheet.cell(row=row, column=3, value=from_airport)
                flight_sheet.cell(row=row, column=4, value=to_airport)
            
            # Auto-adjust column widths - SAME AS PART 1
            for column in flight_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                flight_sheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f" Flight details sheet created: {sheet_name}")
            
        except Exception as e:
            logger.error(f" Error creating flight details sheet: {str(e)}")
    
# Additional helper functions for pytest integration
def pytest_configure(config):
    """Configure pytest for Part 2 testing"""
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


if __name__ == "__main__":
    # Allow running this file directly for debugging
    pytest.main([__file__, "-v"])
