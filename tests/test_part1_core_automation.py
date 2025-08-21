#!/usr/bin/env python3
"""
PART 1: Core Flight Search Automation - Pytest Implementation
EaseMyTrip Flight Search with Filters and Data Extraction

Requirements:
1. Launch the website and perform a flight search using:
    * From location
    * To location
    * Departure Date
2. Apply filters:
    * Only 1 Stop flights
    * Custom Price Range (e.g., ₹4000 to ₹8000)
3. Validate that all listed flights meet the filter criteria.
4. Extract and save the following details for each filtered flight:
    * Airline Name
    * Price
    * From and To Airport Codes
5. Save the extracted data to an Excel file.
6. Consolidate the data:
    * Group by Airline Name
    * Sort by Price
"""

import os
import sys
import pytest
import logging
from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from src.automation.flight_filter_engine import PureUIFilterEngine
from src.utils.config import TestConfig
from src.utils.logger import TestLogger

# Add project root to path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

class TestPart1CoreAutomation:
    """Part 1: Core Flight Search Automation using Pytest Framework"""
    
    # def setup_method(self):
    #     """Setup method called before each test method"""
    #     self.project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    @pytest.mark.part1
    @pytest.mark.ui
    def test_part1_core_automation(self, automation_engine, logger, 
                                  part1_hardcoded_config, results_directory):
        """
        Test Part 1: Core Flight Search Automation
        
        This test performs single automation with hardcoded values as per requirements.
        Tests the complete flow: search → filter → extract → validate → export
        """
        logger.info(" PART 1: Flight Search Automation - Core Implementation (Pytest)")
        logger.info(" Single test case with filters and data extraction")
        logger.info("=" * 80)
        
        # Store test config for reporting
        test_config = part1_hardcoded_config
        
        logger.info(f" Part 1 Configuration:")
        logger.info(f"   Route: {test_config.from_city} -> {test_config.to_city}")
        logger.info(f"   Date: {test_config.departure_date}")
        logger.info(f"   Stops Filter: {test_config.stops_filter}")
        logger.info(f"   Price Range: ₹{test_config.price_min:,} - ₹{test_config.price_max:,}")
        logger.info("-" * 60)
        
        # Execute UI filter testing (which includes all the required steps)
        result = automation_engine.test_ui_filter_functionality(test_config)
        
        # Assert test execution was successful
        assert result is not None, "Automation engine returned None result"
        assert 'status' in result, "Result missing status field"
        
        #  CRITICAL: Check result status FIRST - fail test if core operations failed
        result_status = result.get('status', 'UNKNOWN')
        if result_status == 'FAIL':
            failure_reason = result.get('reason', 'Unknown failure')
            pytest.fail(f" PART 1 TEST FAILED: Core operation failed - {failure_reason}")
        elif result_status == 'ERROR':
            error_details = result.get('error', 'Unknown error')
            pytest.fail(f" PART 1 TEST ERROR: Exception occurred - {error_details}")
        elif result_status != 'SUCCESS':
            pytest.fail(f" PART 1 TEST FAILED: Invalid status '{result_status}' - Expected SUCCESS, FAIL, or ERROR")
        
        # If we reach here, result_status is SUCCESS - proceed with flight analysis
        filter_status = result_status
        flights_found = result.get('before_count', 0)
        
        if flights_found == 0:
            user_message = "  No flights found matching your criteria. Try adjusting filters or different dates."
            logger.info(f" {user_message}")
            filter_status = 'NO_FLIGHTS_FOUND'
        else:
            user_message = f" Success: Found {flights_found} flights for your search"
            logger.info(f" {user_message}")
            filter_status = 'PASSED'
        
        logger.info(f" Filter Status: {filter_status}")
        
        # Extract detailed flight data for Part 1 requirements
        detailed_result = self._extract_and_consolidate_data(result, test_config, logger)
        
        # Store user message for Excel export
        result['user_message'] = user_message
        result['filter_status'] = filter_status
        detailed_result['user_message'] = user_message
        detailed_result['filter_status'] = filter_status
        
        # Assert that data extraction was successful
        assert detailed_result is not None, "Data extraction failed"
        assert 'valid_flights' in detailed_result, "Missing valid_flights in detailed result"
        assert 'filtered_count' in detailed_result, "Missing filtered_count in detailed result"
        
        # Save to Excel with Part 1 specific format
        excel_result = self._save_part1_excel(detailed_result, test_config, logger, results_directory)
        
        # Assert Excel export was successful
        assert excel_result.get('success', False), f"Excel export failed: {excel_result.get('error', 'Unknown error')}"
        assert excel_result.get('filename'), "Excel filename not provided"
        
        # Log results based on flights found
        if flights_found == 0:
            logger.info(" COMPLETED: PART 1 AUTOMATION - NO FLIGHTS FOUND")
        else:
            logger.info(" SUCCESS: PART 1 AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info(f" Total flights found: {flights_found}")
        logger.info(f" Filtered flights (meeting criteria): {detailed_result['filtered_count']}")
        logger.info(f" Excel report: {excel_result['filename']}")
        
        # Data consolidation summary
        if 'consolidation_summary' in detailed_result:
            summary = detailed_result['consolidation_summary']
            logger.info(f" Airlines found: {summary.get('unique_airlines', 0)}")
            logger.info(f" Price range found: ₹{summary.get('min_price', 0):,} - ₹{summary.get('max_price', 0):,}")
        
        # Store results for potential use in other tests or reporting
        self._test_result = {
            'status': 'PASS',
            'flights_found': result.get('before_count', 0),
            'filtered_flights': detailed_result['filtered_count'],
            'excel_file': excel_result['filename'],
            'test_config': test_config.__dict__
        }
        
        logger.info(" PART 1 PYTEST TEST COMPLETED SUCCESSFULLY!")
        print ("=" * 180)
    
    def _extract_and_consolidate_data(self, ui_result: Dict[str, Any], config: TestConfig, logger) -> Dict[str, Any]:
        """Extract detailed flight data and consolidate as per Part 1 requirements"""
        try:
            logger.info(" Extracting and consolidating flight data...")
            
            ui_flights = ui_result.get('ui_filtered_flights', [])
            if not ui_flights:
                logger.warning(" No UI filtered flights to consolidate")
                return {
                    'valid_flights': [],
                    'consolidated_data': [],
                    'filtered_count': 0,
                    'consolidation_summary': {}
                }
            
            # Process each flight for detailed extraction according to assignment requirements
            consolidated_data = []
            valid_flights = []
            airlines_found = set()
            prices_found = []
            
            for i, flight in enumerate(ui_flights, 1):
                try:
                    # Extract only the required fields
                    flight_details = {
                        # Required Fields ONLY:
                        'airline_name': flight.get('airline', 'Unknown'),
                        'price': f"₹{flight.get('price', 0):,}",
                        'price_numeric': flight.get('price', 0),
                        # Use extracted airport codes from website selection
                        'from_airport_code': flight.get('from_code', 'N/A'),
                        'to_airport_code': flight.get('to_code', 'N/A'),
                        
                        # Additional useful fields for consolidation
                        'flight_number': flight.get('flight_number', f'FL{i:03d}'),
                        'stops': flight.get('stops', config.stops_filter),
                    }
                    
                    consolidated_data.append(flight_details)
                    valid_flights.append(flight)
                    
                    # Track for summary
                    if flight_details['airline_name'] != 'Unknown':
                        airlines_found.add(flight_details['airline_name'])
                    if flight_details['price_numeric'] > 0:
                        prices_found.append(flight_details['price_numeric'])
                        
                except Exception as e:
                    logger.warning(f" Error processing flight {i}: {str(e)}")
                    continue
            
            # Sort by price (ascending) as per assignment requirement
            consolidated_data.sort(key=lambda x: x.get('price_numeric', 0))
            
            # Create summary
            summary = {
                'unique_airlines': len(airlines_found),
                'airlines_list': list(airlines_found),
                'min_price': min(prices_found) if prices_found else 0,
                'max_price': max(prices_found) if prices_found else 0,
                'avg_price': sum(prices_found) / len(prices_found) if prices_found else 0,
                'total_flights': len(consolidated_data)
            }
            
            logger.info(f" Consolidation complete: {len(valid_flights)} flights meet all criteria")
            logger.info(f" Airlines found: {', '.join(airlines_found)}")
            logger.info(f" Price range: ₹{summary['min_price']:,} - ₹{summary['max_price']:,}")
            
            return {
                'valid_flights': valid_flights,
                'consolidated_data': consolidated_data,
                'filtered_count': len(valid_flights),
                'consolidation_summary': summary
            }
            
        except Exception as e:
            logger.error(f" Data consolidation error: {str(e)}")
            return {
                'valid_flights': [],
                'consolidated_data': [],
                'filtered_count': 0,
                'consolidation_summary': {}
            }
    
    def _save_part1_excel(self, data_result: Dict[str, Any], config: TestConfig, 
                         logger, results_directory: str) -> Dict[str, Any]:
        """Save consolidated data to Excel with Part 1 specific format and structure"""
        try:
            logger.info(" Saving data to Excel file ")
            
            # Create part1 specific results directory
            part1_results_dir = os.path.join(results_directory, "part1")
            os.makedirs(part1_results_dir, exist_ok=True)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Part1_Core_Automation_Pytest_Results_{timestamp}.xlsx"
            file_path = os.path.join(part1_results_dir, filename)
            
            # Prepare data for Excel export (EXACT Assignment Requirements)
            # For Airline Summary sheet - Group by Airline Name and Sort by Price
            
            # Create Flight Details data - Sort all flights by price (no grouping)
            flight_details_data = []
            all_flights = data_result['consolidated_data'].copy()
            
            # Sort all flights by price for Flight Details sheet
            all_flights.sort(key=lambda x: x.get('price_numeric', 0))
            
            for flight in all_flights:
                excel_row = {
                    'Airline Name': flight.get('airline_name', 'N/A'),
                    'Price': flight.get('price', '₹0'),  # Already formatted with ₹
                    'From Airport Code': flight.get('from_airport_code', 'N/A'),
                    'To Airport Code': flight.get('to_airport_code', 'N/A')
                }
                flight_details_data.append(excel_row)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Sheet 1: Flight Details (Sort by Price Only)
            flight_sheet = workbook.active # Get default sheet
            flight_sheet.title = "Flight Details" # changing the title of sheet
            
            # Headers with styling 
            headers = ['Airline Name', 'Price', 'From Airport Code', 'To Airport Code']
            for col, header in enumerate(headers, 1):
                cell = flight_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows - SORTED BY PRICE (no airline grouping)
            for row, data in enumerate(flight_details_data, 2):
                for col, value in enumerate(data.values(), 1):
                    flight_sheet.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in flight_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                flight_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Airline Summary (APPLY GROUPING LOGIC HERE)
            summary_sheet = workbook.create_sheet("Airline Summary")
            
            # Group by airline for summary sheet
            airline_groups = {}
            for flight in data_result['consolidated_data']:
                airline = flight.get('airline_name', 'Unknown')
                if airline not in airline_groups:
                    airline_groups[airline] = []
                airline_groups[airline].append(flight)
            
            # Summary headers - DETAILED FLIGHT LISTING WITH GROUPING
            summary_headers = ['Airline Name', 'Price', 'From Airport Code', 'To Airport Code']
            for col, header in enumerate(summary_headers, 1):
                cell = summary_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # APPLY GROUPING LOGIC TO AIRLINE SUMMARY SHEET
            # Sort each airline group by price (ascending)
            for airline in airline_groups:
                airline_groups[airline].sort(key=lambda x: x.get('price_numeric', 0))
            
            # Sort airlines by their lowest price flight
            sorted_airlines = sorted(airline_groups.keys(), 
                                   key=lambda airline: min(f.get('price_numeric', 0) for f in airline_groups[airline]))
            
            # Add data with contiguous airline groups and 1-row gaps
            current_row = 2
            for airline_index, airline in enumerate(sorted_airlines):
                # Add a blank row between airline groups (except for the first group)
                if airline_index > 0:
                    current_row += 1  # Skip one row for gap
                
                # Add all flights for this airline (already sorted by price)
                for flight in airline_groups[airline]:
                    summary_sheet.cell(row=current_row, column=1, value=flight.get('airline_name', 'N/A'))
                    summary_sheet.cell(row=current_row, column=2, value=flight.get('price', '₹0'))
                    summary_sheet.cell(row=current_row, column=3, value=flight.get('from_airport_code', 'N/A'))
                    summary_sheet.cell(row=current_row, column=4, value=flight.get('to_airport_code', 'N/A'))
                    current_row += 1
            
            # Auto-adjust summary sheet columns
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(file_path)
            logger.info(f" Excel file saved successfully: {filename}")
            logger.info(f" Data format: Assignment compliant (Airline Name, Price, Airport Codes)")
            logger.info(f" Consolidation: Grouped by airline, sorted by price")
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'sheets': ['Flight Details', 'Airline Summary'],
                'total_flights': len(flight_details_data),
                'airlines_count': len(airline_groups)
            }
            
        except Exception as e:
            logger.error(f" Excel save error: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'filename': None
            }


# Stand-alone execution for PyCharm compatibility
if __name__ == "__main__":
    # This allows running the test directly from PyCharm
    pytest.main([__file__, "-v", "--tb=short"])
